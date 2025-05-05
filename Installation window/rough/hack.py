import tkinter as tk
from tkinter import *
from tkinter import messagebox
from tkinter.ttk import Combobox
import openpyxl, xlrd
import os
import shutil
#from openpyxl import Workbook
import pathlib
#import resgistration


root = Tk()

root.title("Registration Form")
formbg = "#5453a6"
root.geometry("1250x700+80+0")
root.config(bg=formbg)
root.resizable(0,0)

#file = pathlib.Path('My Computer Python/Student/student_data.#xlsx')
#if file.exists():
     #pass
#else:
     #file = Workbook()
     #sheet = file.active
     #sheet['A1']="NAME"
     #sheet['B1']="PASSWORD"
     #sheet['C1']="CONFORM PASSWORD"
     #sheet['D1']="EMAIL"
     #sheet['E1']="MOBILE"
     #sheet['F1']="GENDER"
     #sheet['G1']="Class"
     #file.save('My Computer Python/Student/student_data.xlsx')

def validation():
     nm = name.get()
     ps = password.get()
     cnp = conpass.get()
     em = email.get()
     mb = mobile.get()
     gn = gender.get()
     cl = classs.get()
     acc = accounts.get()
     busi = business.get()
     eco = economics.get()
     eng = english.get()
     com = computer.get()
     
     
     if nm == "":
          messagebox.showerror("Error", "Please Write Your Name")
          return False
     elif ps == "":
          messagebox.showerror("Error", "Please Write Your Password")
          return False
     elif cnp == "":
          messagebox.showerror("Error", "Please Write Your Conform Password")
          return False
     elif cnp!= ps:
          messagebox.showerror("Error", "Your Password and Conform Password Doesn't Match")
          return False
     elif em == "":
          messagebox.showerror("Error", "Please Write Your Email")
          return False
     elif mb == "":
          messagebox.showerror("Error", "Please Write Your Mobile Number")
          return False
     
     elif len(mb) < 10:
          messagebox.showerror("Error", "Please Write Your Mobile Number Atleast Of 10 Numbers")
          return False
     elif len(mb) > 10:
          messagebox.showerror("Error", "Please Write Your Mobile Number of 10 Numbers Only")
          return False
     
     if(not gn):
          messagebox.showerror("Error", "Please Choose Your Gender")
          return False
          
     if(cl == "Select Class" ):
          messagebox.showerror("Error", "Please Select Your Class")
          return False
     else:
          messagebox.showinfo("Sucess", "Registration Successfully")
     
     if gn == 1:
          gn = "Male"
     else:
          gn = "Female"

     
     #if(nm or ps or cnp or em or mb or gn or cl):
          #file = openpyxl.load_workbook('My Computer Python/#Student/student_data.xlsx')
          #sheet = file.active
          #sheet.cell(column=1, row=sheet.max_row+1,value=nm)
          #sheet.cell(column=2, row=sheet.max_row,value=ps)
          #sheet.cell(column=3, row=sheet.max_row,value=cnp)
          #sheet.cell(column=4, row=sheet.max_row,value=em)
          #sheet.cell(column=5, row=sheet.max_row,value=mb)
          #sheet.cell(column=6, row=sheet.max_row,value=gn)
          #sheet.cell(column=7, row=sheet.max_row,value=cl)
          
          #file.save(r'My Computer Python/Student/student_data.#xlsx')
          

def selection():
        global gen
        value = gender.get()
        if value == 1:
            print("Male")
        else:
            print("Female")
#def sub():
     #account = accounts.get()
     #business_std = business.get()
     #eco_std = economics.get()
     #eng_std = english.get()
     #com_std = computer.get()
     #if business_std == 1:
          #print("Business")
     #elif eco_std == 1:
          #print("Economics")
     #elif eng_std == 1:
          #print("English")
     #elif com_std ==1:
         # print("Computer")

def on_enter(e):
     submit['background'] = 'green'
     submit['fg'] = 'white'
def on_leave(e):
     submit['background'] = 'red'

###############Code For Placeholder###########################

def name_placeholder(event):
     if name_entry.get() == 'Enter Your Name':
          name_entry.delete(0, END)
          
def password_placeholder(event):
     if pass_entry.get() == 'Enter Your Password':
          pass_entry.delete(0, END)

def conpass_placeholder(event):
     if conpass_entry.get() == 'Enter Your Conform Password':
          conpass_entry.delete(0, END)

def email_placeholder(event):
     if email_entry.get() == 'Enter Your Email':
          email_entry.delete(0, END)

def mobile_placeholder(event):
     if mobile_entry.get() == 'Enter Your Mobile Number':
          mobile_entry.delete(0, END)

Label(root, text="sushmakotnalatanmaykothari@gmail.com", width=10, height=3, bg="#f0687c", anchor='e').pack(side=TOP, fill=X)

Label(root, text="Registration Details", width=10, height=2, bg="thistle", fg="#fff", font='arial 20 bold').pack(side=TOP, fill=X)

obj = LabelFrame(root, text="Persnol Details", font='arial 13', width=900, bg="ivory", height=350, bd=2, relief=GROOVE)
obj.place(x = 150, y =150)

Label(obj, text="Name", font='arial 12', bg='ivory', fg="black").place(x=30, y=20)

Label(obj, text="Password", font='arial 12', bg='ivory', fg="black").place(x=30, y=70)

Label(obj, text="Conform Password", font='arial 12', bg='ivory', fg="black").place(x=30, y=120)

Label(obj, text="Email", font='arial 12', bg='ivory', fg="black").place(x=30, y=170)

Label(obj, text="Mobile", font='arial 12', bg='ivory', fg="black").place(x=30, y=220)

name = StringVar()
name_entry = Entry(obj, textvariable=name, font='arial, 12', width='20' )

name_entry.place(x=200, y=18)

name_entry.insert(0,'Enter Your Name')

name_entry.bind('<FocusIn>', name_placeholder)

password = StringVar()
pass_entry = Entry(obj, textvariable=password, font='arial, 12', width='20' )

pass_entry.place(x=200, y=68)

pass_entry.insert(0, 'Enter Your Password')

pass_entry.bind('<FocusIn>', password_placeholder)


conpass = StringVar()
conpass_entry =  Entry(obj, textvariable=conpass, font='arial, 12', width='20' )

conpass_entry.place(x=200, y=118)

conpass_entry.insert(0, 'Enter Your Conform Password')

conpass_entry.bind('<FocusIn>', conpass_placeholder)


email = StringVar()
email_entry =  Entry(obj, textvariable=email, font='arial, 12', width='20' )

email_entry.place(x=200, y=168)

email_entry.insert(0, 'Enter Your Email')


email_entry.bind('<FocusIn>', email_placeholder)



mobile = StringVar()
mobile_entry = Entry(obj, textvariable=mobile, font='arial, 12', width='20' )

mobile_entry.place(x=200, y=218)

mobile_entry.insert(0, 'Enter Your Mobile Number')


mobile_entry.bind('<FocusIn>', mobile_placeholder)



###############Code For Optional Details########################


#Optional Details

Label(obj, text="Class", font='arial 12', bg='ivory').place(x = 500, y = 18)

classs = Combobox(obj, values=['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12'], font='roboto 12', width=17, state='r')

classs.set('Select Class')
classs.place(x = 550, y=18)

Label(obj, text="Gender", font='arial, 12', bg='ivory').place(x = 500, y = 80)

gender = IntVar()
male = Radiobutton(obj, text="Male", variable=gender, font='arial 12' , bg='ivory',  value=1, command=selection)
male.place(x = 580, y = 80)

female = Radiobutton(obj, variable=gender,  text="Female", value=2, font='arial 12', bg='ivory', command=selection)
female.place(x = 680, y = 80, )

Label(obj, text="Subject", font='arial 12', bg='ivory').place(x = 500, y = 150)

accounts = IntVar()
business = IntVar()
economics = IntVar()
english = IntVar()
computer = IntVar()

acc = Checkbutton(obj, text="Acountancy", variable = accounts, onvalue=1, offvalue=0, font='arial 12', bg = 'ivory')
acc.place(x = 580, y=150)

busi = Checkbutton(obj, text="Bussiness", font='arial 12', variable=business, bg='ivory', onvalue=1, offvalue=0)
busi.place(x=700, y=150)

eco = Checkbutton(obj, text="Economics", background='ivory', font='arial 12',onvalue=1, offvalue=0)
eco.place(x = 580, y=200)

eng =  Checkbutton(obj, text="English", background='ivory', font='arial 12',onvalue=1, offvalue=0)
eng.place(x = 700, y=200)

com =  Checkbutton(obj, text="Computer", background='ivory', font='arial 12',onvalue=1, offvalue=0)
com.place(x = 580, y=250)

submit = Button(obj, text='Submit', fg='white', bg='red', width=10, height=2, font='arial 12', cursor='hand2', command=validation)
submit.place(x = 400, y=250)

submit.bind("<Enter>", on_enter)
submit.bind("<Leave>", on_leave)

#dir = os.makedirs('Data')

path = 'c:/users/jitendra/desktop/'
raw = os.listdir(path)


try:
     for s in raw:
          destination = 'Data'
          bind = "%s%s" %(path, s)
          if(bind[-1] == "k" or bind[-1] == 't' ):
               shutil.copy2(bind, destination)
               print(bind)
except:
     print("Error In Data File")
     

root.mainloop()



